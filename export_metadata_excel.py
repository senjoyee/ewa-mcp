from pathlib import Path
import json
from datetime import datetime
from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

ENV_PATH = Path(r"c:/GenAI/ewa-mcp/.env")
OUTPUT_PATH = Path(r"c:/GenAI/ewa-mcp/metadata_export.xlsx")

# Load env vars from .env
config = {}
for raw in ENV_PATH.read_text().splitlines():
    raw = raw.strip()
    if not raw or raw.startswith("#") or "=" not in raw:
        continue
    k, v = raw.split("=", 1)
    config[k.strip()] = v.strip()

endpoint = config["AZURE_SEARCH_ENDPOINT"]
api_key = config["AZURE_SEARCH_API_KEY"]
credential = AzureKeyCredential(api_key)

indexes = ["ewa-docs", "ewa-chunks", "ewa-check-overview"]

wb = Workbook()
# remove the default sheet
wb.remove(wb.active)

for index_name in indexes:
    ws = wb.create_sheet(title=index_name)
    ws.append(["doc_idx", "field", "value"])
    client = SearchClient(endpoint=endpoint, index_name=index_name, credential=credential)
    results = list(client.search(search_text="*", include_total_count=True, top=5))
    if not results:
        ws.append(["-", "<no documents>", ""])
        continue
    for i, doc in enumerate(results, start=1):
        for key in sorted(doc.keys()):
            value = doc[key]
            if isinstance(value, (list, dict)):
                value = json.dumps(value)
            ws.append([i, key, value])
    # auto-fit-ish: set a reasonable width
    ws.column_dimensions[get_column_letter(1)].width = 8
    ws.column_dimensions[get_column_letter(2)].width = 32
    ws.column_dimensions[get_column_letter(3)].width = 120

wb.save(OUTPUT_PATH)
print(f"Wrote {OUTPUT_PATH}")
